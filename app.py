from flask import Flask, render_template, request, redirect, url_for, jsonify, flash, send_from_directory
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from flask_mail import Mail, Message
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from datetime import datetime, timedelta
import os
import random
import string
import time
import requests

app = Flask(__name__)
app.config['SECRET_KEY'] = 'menstrual-tracker-secret-key-2024'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///menstrual.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# 邮件配置
app.config['MAIL_SERVER'] = 'smtp.163.com'
app.config['MAIL_PORT'] = 465
app.config['MAIL_USE_SSL'] = True
app.config['MAIL_USERNAME'] = 'yzh189983@163.com'
app.config['MAIL_PASSWORD'] = 'yzh18998301631'
app.config['MAIL_DEFAULT_SENDER'] = 'yzh189983@163.com'

# 文件上传配置
app.config['UPLOAD_FOLDER'] = 'static/uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max
app.config['AVATAR_FOLDER'] = 'static/avatars'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}

# DeepSeek API 配置
DEEPSEEK_API_KEY = "sk-438f7ee6c06f4f75b0eca2a7bb6106fa"
DEEPSEEK_API_URL = "https://api.deepseek.com/v1/chat/completions"

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['AVATAR_FOLDER'], exist_ok=True)

db = SQLAlchemy(app)
mail = Mail(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

# 验证码存储 (临时存储)
verification_codes = {}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# 用户模型
class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(200), nullable=False)
    email = db.Column(db.String(120), unique=True)
    nickname = db.Column(db.String(80))
    phone = db.Column(db.String(20))
    birthday = db.Column(db.Date)
    avatar = db.Column(db.String(200))
    is_admin = db.Column(db.Boolean, default=False)  # 管理员标识
    partner_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)  # 情侣绑定
    created_at = db.Column(db.DateTime, default=datetime.now)
    periods = db.relationship('Period', backref='user', lazy=True)
    partner = db.relationship('User', remote_side=[id], backref='partner_of_me')

# 月经记录模型
class Period(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    start_date = db.Column(db.Date, nullable=False)
    end_date = db.Column(db.Date, nullable=False)
    flow = db.Column(db.String(20), default='medium')
    pain_level = db.Column(db.Integer, default=0)  # 疼痛等级 0-10
    symptoms = db.Column(db.String(200))  # 症状，多个用逗号分隔
    notes = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.now)

# 学习记录模型
class StudyRecord(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    user = db.relationship('User', backref='study_records')
    date = db.Column(db.Date, nullable=False)
    subject = db.Column(db.String(100), nullable=False)  # 学习科目
    duration = db.Column(db.Integer, nullable=False)      # 学习时长（分钟）
    plan = db.Column(db.Text)                             # 学习计划/内容
    notes = db.Column(db.Text)                             # 备注
    created_at = db.Column(db.DateTime, default=datetime.now)

# 工作记录模型
class WorkRecord(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    user = db.relationship('User', backref='work_records')
    date = db.Column(db.Date, nullable=False)
    task = db.Column(db.String(200))                      # 工作任务
    task_duration = db.Column(db.Integer, default=0)      # 任务时长（分钟）
    work_start = db.Column(db.Time)                       # 上班时间
    work_end = db.Column(db.Time)                         # 下班时间
    overtime = db.Column(db.Integer, default=0)           # 加班时长（分钟）
    notes = db.Column(db.Text)                            # 备注
    created_at = db.Column(db.DateTime, default=datetime.now)

# 用户反馈模型
class Feedback(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    user = db.relationship('User', backref='feedbacks')
    feedback_type = db.Column(db.String(20), nullable=False)  # suggest/bug/praise/other
    contact = db.Column(db.String(100))                        # 联系方式
    content = db.Column(db.Text, nullable=False)               # 反馈内容
    reply = db.Column(db.Text)                                 # 回复内容
    status = db.Column(db.String(20), default='pending')       # pending/replied
    created_at = db.Column(db.DateTime, default=datetime.now)

# 好友关系模型
class Friend(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    friend_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    status = db.Column(db.String(20), default='pending')  # pending/accepted/rejected
    relation_type = db.Column(db.String(20), default='friend')  # friend/partner
    created_at = db.Column(db.DateTime, default=datetime.now)
    
    user = db.relationship('User', foreign_keys=[user_id], backref='friends')
    friend = db.relationship('User', foreign_keys=[friend_id], backref='friend_of')

# 聊天消息模型
class ChatMessage(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    sender_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    receiver_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    message = db.Column(db.Text, nullable=False)
    message_type = db.Column(db.String(20), default='text')  # text/record
    record_id = db.Column(db.Integer)  # 分享的记录ID
    record_type = db.Column(db.String(20))  # period/study/work
    is_read = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.now)
    
    sender = db.relationship('User', foreign_keys=[sender_id], backref='sent_messages')
    receiver = db.relationship('User', foreign_keys=[receiver_id], backref='received_messages')

# 错题集模型
class WrongQuestion(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    subject = db.Column(db.String(50))  # 学科
    question = db.Column(db.Text, nullable=False)  # 错题题目
    user_answer = db.Column(db.Text)  # 用户的答案
    ai_explanation = db.Column(db.Text)  # AI的讲解
    knowledge_points = db.Column(db.Text)  # 知识点
    practice_question = db.Column(db.Text)  # 练习题
    created_at = db.Column(db.DateTime, default=datetime.now)
    
    user = db.relationship('User', backref='wrong_questions')

# ========== 社区模块 ==========

# 帖子模型
class Post(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    content = db.Column(db.Text, nullable=False)
    images = db.Column(db.String(500))  # 多图用逗号分隔
    topic = db.Column(db.String(50), default='general')  # 话题标签
    is_anonymous = db.Column(db.Boolean, default=False)
    likes_count = db.Column(db.Integer, default=0)
    comments_count = db.Column(db.Integer, default=0)
    created_at = db.Column(db.DateTime, default=datetime.now)
    
    user = db.relationship('User', backref='posts')

# 帖子点赞
class PostLike(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    post_id = db.Column(db.Integer, db.ForeignKey('post.id'), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.now)
    
    post = db.relationship('Post', backref='likes')

# 帖子评论
class PostComment(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    post_id = db.Column(db.Integer, db.ForeignKey('post.id'), nullable=False)
    content = db.Column(db.Text, nullable=False)
    parent_id = db.Column(db.Integer, db.ForeignKey('post_comment.id'), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.now)
    
    user = db.relationship('User', backref='post_comments')
    post = db.relationship('Post', backref='comments')
    parent = db.relationship('PostComment', remote_side=[id], backref='replies')

# 用户关注
class UserFollow(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    follower_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    following_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.now)
    
    follower = db.relationship('User', foreign_keys=[follower_id], backref='following')
    following = db.relationship('User', foreign_keys=[following_id], backref='followers')

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

# 首页 - 显示日历
@app.route('/')
@login_required
def index():
    periods = Period.query.filter_by(user_id=current_user.id).order_by(Period.start_date.desc()).all()
    return render_template('calendar.html', periods=periods, user=current_user, timedelta=timedelta, datetime=datetime)

# 学习记录页面
@app.route('/study')
@login_required
def study():
    records = StudyRecord.query.filter_by(user_id=current_user.id).order_by(StudyRecord.date.desc()).all()
    return render_template('study.html', records=records, user=current_user, timedelta=timedelta)

# 添加学习记录
@app.route('/study/add', methods=['POST'])
@login_required
def add_study():
    date = datetime.strptime(request.form.get('date'), '%Y-%m-%d').date()
    subject = request.form.get('subject')
    duration = int(request.form.get('duration'))
    plan = request.form.get('plan', '')
    notes = request.form.get('notes', '')
    
    record = StudyRecord(
        user_id=current_user.id,
        date=date,
        subject=subject,
        duration=duration,
        plan=plan,
        notes=notes
    )
    db.session.add(record)
    db.session.commit()
    
    flash('学习记录添加成功！', 'success')
    return redirect(url_for('study'))

# 删除学习记录
@app.route('/study/delete/<int:record_id>')
@login_required
def delete_study(record_id):
    record = StudyRecord.query.get_or_404(record_id)
    if record.user_id == current_user.id:
        db.session.delete(record)
        db.session.commit()
        flash('学习记录已删除', 'success')
    return redirect(url_for('study'))

# 编辑学习记录
@app.route('/study/edit/<int:record_id>', methods=['POST'])
@login_required
def edit_study(record_id):
    record = StudyRecord.query.get_or_404(record_id)
    if record.user_id != current_user.id:
        return jsonify({'success': False, 'message': '无权修改'})
    
    record.subject = request.form.get('subject', record.subject)
    record.duration = int(request.form.get('duration', record.duration))
    record.date = datetime.strptime(request.form.get('date'), '%Y-%m-%d').date()
    record.plan = request.form.get('plan', record.plan)
    record.notes = request.form.get('notes', record.notes)
    
    db.session.commit()
    return jsonify({'success': True, 'message': '更新成功'})

# 批量删除学习记录
@app.route('/study/batch_delete', methods=['POST'])
@login_required
def batch_delete_study():
    ids = request.form.get('ids', '').split(',')
    deleted_count = 0
    for id_str in ids:
        if id_str.strip():
            try:
                rid = int(id_str.strip())
                record = StudyRecord.query.filter_by(id=rid, user_id=current_user.id).first()
                if record:
                    db.session.delete(record)
                    deleted_count += 1
            except:
                pass
    db.session.commit()
    flash(f'已删除 {deleted_count} 条学习记录', 'success')
    return redirect(url_for('study'))

# API: 获取学习记录数据
@app.route('/api/study_data')
@login_required
def api_study_data():
    records = StudyRecord.query.filter_by(user_id=current_user.id).all()
    data = []
    for r in records:
        data.append({
            'id': r.id,
            'date': r.date.strftime('%Y-%m-%d'),
            'subject': r.subject,
            'duration': r.duration,
            'plan': r.plan,
            'notes': r.notes
        })
    return jsonify(data)

# 错题集页面
@app.route('/wrong_questions')
@login_required
def wrong_questions():
    subject_filter = request.args.get('subject', '')
    query = WrongQuestion.query.filter_by(user_id=current_user.id)
    if subject_filter:
        query = query.filter_by(subject=subject_filter)
    questions = query.order_by(WrongQuestion.created_at.desc()).all()
    
    # 获取所有学科用于筛选
    all_subjects = db.session.query(WrongQuestion.subject).filter_by(user_id=current_user.id).distinct().all()
    subjects = [s[0] for s in all_subjects if s[0]]
    
    return render_template('wrong_questions.html', 
                           questions=questions, 
                           user=current_user,
                           subject_filter=subject_filter,
                           subjects=subjects)

# 删除错题
@app.route('/wrong_questions/delete/<int:qid>', methods=['POST'])
@login_required
def delete_wrong_question(qid):
    question = WrongQuestion.query.get_or_404(qid)
    if question.user_id == current_user.id:
        db.session.delete(question)
        db.session.commit()
        flash('错题已删除', 'success')
    return redirect(url_for('wrong_questions'))

# 批量删除错题
@app.route('/wrong_questions/batch_delete', methods=['POST'])
@login_required
def batch_delete_wrong_questions():
    ids = request.form.get('ids', '').split(',')
    deleted_count = 0
    for id_str in ids:
        if id_str.strip():
            try:
                qid = int(id_str.strip())
                question = WrongQuestion.query.filter_by(id=qid, user_id=current_user.id).first()
                if question:
                    db.session.delete(question)
                    deleted_count += 1
            except:
                pass
    db.session.commit()
    flash(f'已删除 {deleted_count} 道错题', 'success')
    return redirect(url_for('wrong_questions'))

# 编辑错题
@app.route('/wrong_questions/edit/<int:qid>', methods=['POST'])
@login_required
def edit_wrong_question(qid):
    question = WrongQuestion.query.get_or_404(qid)
    if question.user_id != current_user.id:
        return jsonify({'success': False, 'message': '无权修改'})
    
    question.subject = request.form.get('subject', question.subject)
    question.question = request.form.get('question', question.question)
    question.user_answer = request.form.get('user_answer', question.user_answer)
    question.ai_explanation = request.form.get('ai_explanation', question.ai_explanation)
    question.knowledge_points = request.form.get('knowledge_points', question.knowledge_points)
    question.practice_question = request.form.get('practice_question', question.practice_question)
    
    db.session.commit()
    return jsonify({'success': True, 'message': '更新成功'})

# 工作记录页面
@app.route('/work')
@login_required
def work():
    records = WorkRecord.query.filter_by(user_id=current_user.id).order_by(WorkRecord.date.desc()).all()
    return render_template('work.html', records=records, user=current_user, timedelta=timedelta)

# 添加工作记录
@app.route('/work/add', methods=['POST'])
@login_required
def add_work():
    date = datetime.strptime(request.form.get('date'), '%Y-%m-%d').date()
    task = request.form.get('task', '')
    task_duration = int(request.form.get('task_duration', 0))
    work_start = datetime.strptime(request.form.get('work_start'), '%H:%M').time() if request.form.get('work_start') else None
    work_end = datetime.strptime(request.form.get('work_end'), '%H:%M').time() if request.form.get('work_end') else None
    overtime = int(request.form.get('overtime', 0))
    notes = request.form.get('notes', '')
    
    record = WorkRecord(
        user_id=current_user.id,
        date=date,
        task=task,
        task_duration=task_duration,
        work_start=work_start,
        work_end=work_end,
        overtime=overtime,
        notes=notes
    )
    db.session.add(record)
    db.session.commit()
    
    flash('工作记录添加成功！', 'success')
    return redirect(url_for('work'))

# 删除工作记录
@app.route('/work/delete/<int:record_id>')
@login_required
def delete_work(record_id):
    record = WorkRecord.query.get_or_404(record_id)
    if record.user_id == current_user.id:
        db.session.delete(record)
        db.session.commit()
        flash('工作记录已删除', 'success')
    return redirect(url_for('work'))

# 编辑工作记录
@app.route('/work/edit/<int:record_id>', methods=['POST'])
@login_required
def edit_work(record_id):
    record = WorkRecord.query.get_or_404(record_id)
    if record.user_id != current_user.id:
        return jsonify({'success': False, 'message': '无权修改'})
    
    record.task = request.form.get('task', record.task)
    record.task_duration = int(request.form.get('task_duration', record.task_duration))
    record.work_start = request.form.get('work_start') or None
    record.work_end = request.form.get('work_end') or None
    if record.work_start and record.work_end:
        try:
            start = datetime.strptime(record.work_start, '%H:%M')
            end = datetime.strptime(record.work_end, '%H:%M')
            record.overtime = max(0, int((end - start).total_seconds() / 60) - 480)
        except:
            pass
    record.notes = request.form.get('notes', record.notes)
    
    db.session.commit()
    return jsonify({'success': True, 'message': '更新成功'})

# 批量删除工作记录
@app.route('/work/batch_delete', methods=['POST'])
@login_required
def batch_delete_work():
    ids = request.form.get('ids', '').split(',')
    deleted_count = 0
    for id_str in ids:
        if id_str.strip():
            try:
                rid = int(id_str.strip())
                record = WorkRecord.query.filter_by(id=rid, user_id=current_user.id).first()
                if record:
                    db.session.delete(record)
                    deleted_count += 1
            except:
                pass
    db.session.commit()
    flash(f'已删除 {deleted_count} 条工作记录', 'success')
    return redirect(url_for('work'))

# 总日历页面 - 查看所有记录
@app.route('/calendar')
@login_required
def all_calendar():
    periods = Period.query.filter_by(user_id=current_user.id).all()
    study_records = StudyRecord.query.filter_by(user_id=current_user.id).all()
    work_records = WorkRecord.query.filter_by(user_id=current_user.id).all()
    return render_template('all_calendar.html', 
                           periods=periods, 
                           study_records=study_records,
                           work_records=work_records,
                           user=current_user, 
                           timedelta=timedelta)

# 用户反馈页面
@app.route('/feedback', methods=['GET', 'POST'])
@login_required
def feedback():
    if request.method == 'POST':
        feedback_type = request.form.get('feedback_type')
        contact = request.form.get('contact', '')
        content = request.form.get('content')
        
        if not content:
            flash('请填写反馈内容！', 'error')
            return redirect(url_for('feedback'))
        
        feedback = Feedback(
            user_id=current_user.id,
            feedback_type=feedback_type,
            contact=contact,
            content=content,
            status='pending'
        )
        db.session.add(feedback)
        db.session.commit()
        
        flash('感谢您的反馈！我们会尽快处理！', 'success')
        return redirect(url_for('feedback'))
    
    # 获取当前用户的反馈记录
    feedbacks = Feedback.query.filter_by(user_id=current_user.id).order_by(Feedback.created_at.desc()).all()
    return render_template('feedback.html', feedbacks=feedbacks, user=current_user)

# 聊天主页 - 好友列表
@app.route('/chat')
@login_required
def chat():
    # 获取好友列表（已接受的）
    friends = db.session.query(User).join(
        Friend, (Friend.friend_id == User.id) | (Friend.user_id == User.id)
    ).filter(
        ((Friend.user_id == current_user.id) | (Friend.friend_id == current_user.id)),
        Friend.status == 'accepted',
        User.id != current_user.id
    ).all()
    
    # 获取未读消息数
    unread_count = ChatMessage.query.filter_by(receiver_id=current_user.id, is_read=False).count()
    
    # 获取好友请求
    friend_requests = db.session.query(User, Friend).join(
        Friend, Friend.user_id == User.id
    ).filter(
        Friend.friend_id == current_user.id,
        Friend.status == 'pending'
    ).all()
    
    return render_template('chat.html', friends=friends, friend_requests=friend_requests, 
                           unread_count=unread_count, user=current_user)

# 添加好友
@app.route('/chat/add_friend', methods=['POST'])
@login_required
def add_friend():
    username = request.form.get('username')
    if not username:
        flash('请输入用户名！', 'error')
        return redirect(url_for('chat'))
    
    # 查找用户
    friend_user = User.query.filter_by(username=username).first()
    if not friend_user:
        flash('用户不存在！', 'error')
        return redirect(url_for('chat'))
    
    if friend_user.id == current_user.id:
        flash('不能添加自己为好友！', 'error')
        return redirect(url_for('chat'))
    
    # 检查是否已经是好友
    existing = Friend.query.filter(
        ((Friend.user_id == current_user.id) & (Friend.friend_id == friend_user.id)) |
        ((Friend.user_id == friend_user.id) & (Friend.friend_id == current_user.id))
    ).first()
    
    if existing:
        if existing.status == 'accepted':
            flash('你们已经是好友了！', 'info')
        else:
            flash('已经发送过好友请求了！', 'info')
        return redirect(url_for('chat'))
    
    # 创建好友请求
    friend = Friend(user_id=current_user.id, friend_id=friend_user.id, status='pending')
    db.session.add(friend)
    db.session.commit()
    
    flash(f'已向 {username} 发送好友请求！', 'success')
    return redirect(url_for('chat'))

# 同意/拒绝好友请求
@app.route('/chat/friend_request/<int:request_id>/<action>')
@login_required
def friend_request(request_id, action):
    friend = Friend.query.get_or_404(request_id)
    
    if friend.friend_id != current_user.id:
        flash('无权操作！', 'error')
        return redirect(url_for('chat'))
    
    if action == 'accept':
        friend.status = 'accepted'
        db.session.commit()
        flash('已同意好友请求！', 'success')
    elif action == 'reject':
        db.session.delete(friend)
        db.session.commit()
        flash('已拒绝好友请求！', 'success')
    
    return redirect(url_for('chat'))

# 与好友聊天页面
@app.route('/chat/<int:friend_id>')
@login_required
def chat_room(friend_id):
    # 检查是否是好友
    friendship = Friend.query.filter(
        ((Friend.user_id == current_user.id) & (Friend.friend_id == friend_id)) |
        ((Friend.user_id == friend_id) & (Friend.friend_id == current_user.id))
    ).filter(Friend.status == 'accepted').first()
    
    if not friendship:
        flash('你们还不是好友！', 'error')
        return redirect(url_for('chat'))
    
    friend = User.query.get_or_404(friend_id)
    
    # 获取聊天记录
    messages = ChatMessage.query.filter(
        ((ChatMessage.sender_id == current_user.id) & (ChatMessage.receiver_id == friend_id)) |
        ((ChatMessage.sender_id == friend_id) & (ChatMessage.receiver_id == current_user.id))
    ).order_by(ChatMessage.created_at.asc()).all()
    
    # 标记未读消息为已读
    ChatMessage.query.filter_by(sender_id=friend_id, receiver_id=current_user.id, is_read=False).update({'is_read': True})
    db.session.commit()
    
    # 获取所有记录用于分享（自己的记录）
    periods = Period.query.filter_by(user_id=current_user.id).all()
    study_records = StudyRecord.query.filter_by(user_id=current_user.id).all()
    work_records = WorkRecord.query.filter_by(user_id=current_user.id).all()
    
    # 获取好友的记录（用于显示对方分享的记录详情）
    friend_periods = Period.query.filter_by(user_id=friend_id).all()
    friend_study_records = StudyRecord.query.filter_by(user_id=friend_id).all()
    friend_work_records = WorkRecord.query.filter_by(user_id=friend_id).all()
    
    return render_template('chat_room.html', friend=friend, messages=messages,
                           periods=periods, study_records=study_records, 
                           work_records=work_records,
                           friend_periods=friend_periods,
                           friend_study_records=friend_study_records,
                           friend_work_records=friend_work_records,
                           user=current_user)

# 发送消息
@app.route('/chat/send', methods=['POST'])
@login_required
def send_message():
    receiver_id = request.form.get('receiver_id')
    message = request.form.get('message', '').strip()
    message_type = request.form.get('message_type', 'text')
    record_id = request.form.get('record_id')
    record_type = request.form.get('record_type')
    
    if not receiver_id:
        flash('请选择接收者！', 'error')
        return redirect(url_for('chat'))
    
    receiver_id = int(receiver_id)
    
    # 检查是否是好友
    friendship = Friend.query.filter(
        ((Friend.user_id == current_user.id) & (Friend.friend_id == receiver_id)) |
        ((Friend.user_id == receiver_id) & (Friend.friend_id == current_user.id))
    ).filter(Friend.status == 'accepted').first()
    
    if not friendship:
        flash('你们还不是好友！', 'error')
        return redirect(url_for('chat'))
    
    if message_type == 'text' and not message:
        flash('消息内容不能为空！', 'error')
        return redirect(url_for('chat_room', friend_id=receiver_id))
    
    # 创建消息
    chat_msg = ChatMessage(
        sender_id=current_user.id,
        receiver_id=receiver_id,
        message=message,
        message_type=message_type,
        record_id=record_id,
        record_type=record_type
    )
    db.session.add(chat_msg)
    db.session.commit()
    
    return redirect(url_for('chat_room', friend_id=receiver_id))

# 注册
@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        if not username or not password:
            flash('请填写用户名和密码！', 'error')
            return redirect(url_for('register'))
        
        if User.query.filter_by(username=username).first():
            flash('用户名已存在！', 'error')
            return redirect(url_for('register'))
        
        user = User(
            username=username, 
            password_hash=generate_password_hash(password)
        )
        db.session.add(user)
        db.session.commit()
        
        login_user(user)
        return redirect(url_for('index'))
    
    return render_template('register.html')

# 登录
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        user = User.query.filter_by(username=username).first()
        if user and check_password_hash(user.password_hash, password):
            login_user(user)
            return redirect(url_for('index'))
        
        flash('用户名或密码错误！', 'error')
    return render_template('login.html')

# 登出
@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

# 管理员登录
@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        # 验证管理员账号
        if username == 'admin' and password == 'yzh18998301631':
            # 查找或创建管理员用户
            admin = User.query.filter_by(username='admin').first()
            if not admin:
                admin = User(
                    username='admin',
                    password_hash=generate_password_hash('yzh18998301631'),
                    is_admin=True
                )
                db.session.add(admin)
                db.session.commit()
            else:
                admin.is_admin = True
                db.session.commit()
            
            login_user(admin)
            return redirect(url_for('admin_dashboard'))
        
        flash('管理员账号或密码错误！', 'error')
    return render_template('admin_login.html')

# 管理员后台
@app.route('/admin')
@login_required
def admin_dashboard():
    # 检查是否是管理员
    if not current_user.is_admin:
        flash('您没有权限访问该页面！', 'error')
        return redirect(url_for('index'))
    
    # 获取所有用户
    users = User.query.filter_by(is_admin=False).all()
    
    # 获取所有记录
    all_periods = Period.query.all()
    all_study = StudyRecord.query.all()
    all_work = WorkRecord.query.all()
    all_feedback = Feedback.query.order_by(Feedback.created_at.desc()).all()
    
    return render_template('admin.html', 
                           users=users,
                           all_periods=all_periods,
                           all_study=all_study,
                           all_work=all_work,
                           all_feedback=all_feedback,
                           user=current_user)

# 回复用户反馈
@app.route('/admin/feedback/reply/<int:feedback_id>', methods=['POST'])
@login_required
def admin_feedback_reply(feedback_id):
    if not current_user.is_admin:
        flash('您没有权限操作！', 'error')
        return redirect(url_for('index'))
    
    feedback = Feedback.query.get_or_404(feedback_id)
    feedback.reply = request.form.get('reply')
    feedback.status = 'replied'
    db.session.commit()
    
    flash('回复成功！', 'success')
    return redirect(url_for('admin_dashboard'))

# 删除用户
@app.route('/admin/delete_user', methods=['POST'])
@login_required
def admin_delete_user():
    if not current_user.is_admin:
        flash('您没有权限操作！', 'error')
        return redirect(url_for('index'))
    
    user_id = request.form.get('user_id')
    user = User.query.get_or_404(user_id)
    
    # 不能删除自己
    if user.id == current_user.id:
        flash('不能删除自己的账号！', 'error')
        return redirect(url_for('admin_dashboard'))
    
    # 删除用户的所有记录
    Period.query.filter_by(user_id=user.id).delete()
    StudyRecord.query.filter_by(user_id=user.id).delete()
    WorkRecord.query.filter_by(user_id=user.id).delete()
    Feedback.query.filter_by(user_id=user.id).delete()
    
    # 删除用户
    db.session.delete(user)
    db.session.commit()
    
    flash(f'用户 {user.username} 已删除！', 'success')
    return redirect(url_for('admin_dashboard'))

# 重置用户密码
@app.route('/admin/reset_password', methods=['POST'])
@login_required
def admin_reset_password():
    if not current_user.is_admin:
        flash('您没有权限操作！', 'error')
        return redirect(url_for('index'))
    
    user_id = request.form.get('user_id')
    new_password = request.form.get('new_password')
    confirm_password = request.form.get('confirm_password')
    
    if new_password != confirm_password:
        flash('两次输入的密码不一致！', 'error')
        return redirect(url_for('admin_dashboard'))
    
    user = User.query.get_or_404(user_id)
    user.password_hash = generate_password_hash(new_password)
    db.session.commit()
    
    flash(f'用户 {user.username} 的密码已重置！', 'success')
    return redirect(url_for('admin_dashboard'))

# 忘记密码页面
@app.route('/forgot', methods=['GET', 'POST'])
def forgot_password():
    if request.method == 'POST':
        action = request.form.get('action')
        
        if action == 'send_code':
            # 发送验证码
            email = request.form.get('email')
            user = User.query.filter_by(email=email).first()
            
            if not user:
                flash('该邮箱未注册！', 'error')
            else:
                # 生成6位验证码
                code = ''.join(random.choices(string.digits, k=6))
                verification_codes[email] = {
                    'code': code,
                    'time': time.time(),
                    'user_id': user.id
                }
                
                try:
                    msg = Message('🌸 月经记录本 - 验证码', recipients=[email])
                    msg.body = f'''您好 {user.username}！

您的验证码是：{code}

验证码有效期为10分钟，请尽快完成验证。

如果这不是您的操作，请忽略此邮件。

---
月经记录本'''
                    mail.send(msg)
                    flash('验证码已发送到您的邮箱！', 'success')
                except Exception as e:
                    flash(f'发送失败: {str(e)}', 'error')
            
            return render_template('forgot.html', email=email, step='verify')
        
        elif action == 'reset':
            # 重置密码
            email = request.form.get('email')
            code = request.form.get('code')
            new_password = request.form.get('new_password')
            confirm_password = request.form.get('confirm_password')
            
            if new_password != confirm_password:
                flash('两次输入的密码不一致！', 'error')
                return render_template('forgot.html', email=email, step='reset')
            
            # 验证验证码
            if email not in verification_codes:
                flash('请先获取验证码！', 'error')
                return render_template('forgot.html', email=email, step='verify')
            
            if time.time() - verification_codes[email]['time'] > 600:  # 10分钟有效期
                del verification_codes[email]
                flash('验证码已过期，请重新获取！', 'error')
                return render_template('forgot.html', email=email, step='verify')
            
            if verification_codes[email]['code'] != code:
                flash('验证码错误！', 'error')
                return render_template('forgot.html', email=email, step='reset')
            
            # 更新密码
            user = User.query.filter_by(email=email).first()
            user.password_hash = generate_password_hash(new_password)
            db.session.commit()
            
            del verification_codes[email]
            flash('密码重置成功！请使用新密码登录', 'success')
            return redirect(url_for('login'))
    
    return render_template('forgot.html')

# 个人信息页面
@app.route('/profile', methods=['GET', 'POST'])
@login_required
def profile():
    if request.method == 'POST':
        # 处理表单提交
        current_user.nickname = request.form.get('nickname', '')
        current_user.phone = request.form.get('phone', '')
        
        # 处理邮箱修改
        new_email = request.form.get('email', '')
        if new_email and new_email != current_user.email:
            # 检查邮箱是否已被占用
            existing = User.query.filter_by(email=new_email).first()
            if existing and existing.id != current_user.id:
                flash('该邮箱已被其他用户使用！', 'error')
                return redirect(url_for('profile'))
            current_user.email = new_email
        
        birthday_str = request.form.get('birthday')
        if birthday_str:
            try:
                current_user.birthday = datetime.strptime(birthday_str, '%Y-%m-%d').date()
            except:
                pass
        
        # 处理头像上传
        if 'avatar' in request.files:
            file = request.files['avatar']
            if file and allowed_file(file.filename):
                filename = secure_filename(f"avatar_{current_user.id}_{int(time.time())}.{file.filename.rsplit('.', 1)[1].lower()}")
                file.save(os.path.join(app.config['AVATAR_FOLDER'], filename))
                
                # 删除旧头像
                if current_user.avatar:
                    old_path = os.path.join(app.config['AVATAR_FOLDER'], current_user.avatar)
                    if os.path.exists(old_path):
                        os.remove(old_path)
                
                current_user.avatar = filename
        
        db.session.commit()
        flash('个人信息更新成功！', 'success')
        return redirect(url_for('profile'))
    
    # 获取待处理的情侣请求
    partner_requests = Friend.query.filter_by(
        friend_id=current_user.id,
        relation_type='partner',
        status='pending'
    ).all()
    
    return render_template('profile.html', user=current_user, partner_requests=partner_requests)

# 头像文件访问
@app.route('/avatars/<filename>')
def avatars(filename):
    return send_from_directory(app.config['AVATAR_FOLDER'], filename)

# 添加记录
@app.route('/add', methods=['POST'])
@login_required
def add_period():
    start_date = datetime.strptime(request.form.get('start_date'), '%Y-%m-%d').date()
    end_date = datetime.strptime(request.form.get('end_date'), '%Y-%m-%d').date()
    flow = request.form.get('flow', 'medium')
    pain_level = int(request.form.get('pain_level', 0))
    # 多个checkbox用getlist获取，用逗号分隔
    symptoms_list = request.form.getlist('symptoms')
    symptoms = ','.join(symptoms_list) if symptoms_list else ''
    notes = request.form.get('notes', '')
    
    period = Period(
        user_id=current_user.id,
        start_date=start_date,
        end_date=end_date,
        flow=flow,
        pain_level=pain_level,
        symptoms=symptoms,
        notes=notes
    )
    db.session.add(period)
    db.session.commit()
    
    flash('记录添加成功！', 'success')
    return redirect(url_for('index'))

# 删除记录
@app.route('/delete/<int:period_id>')
@login_required
def delete_period(period_id):
    period = Period.query.get_or_404(period_id)
    if period.user_id == current_user.id:
        db.session.delete(period)
        db.session.commit()
        flash('记录已删除', 'success')
    return redirect(url_for('index'))

# 编辑记录
@app.route('/edit/<int:period_id>', methods=['POST'])
@login_required
def edit_period(period_id):
    period = Period.query.get_or_404(period_id)
    if period.user_id != current_user.id:
        return jsonify({'success': False, 'message': '无权修改'})
    
    period.start_date = datetime.strptime(request.form.get('start_date'), '%Y-%m-%d').date()
    period.end_date = datetime.strptime(request.form.get('end_date'), '%Y-%m-%d').date()
    period.flow = request.form.get('flow', period.flow)
    period.pain_level = int(request.form.get('pain_level', period.pain_level or 0))
    period.symptoms = request.form.get('symptoms', period.symptoms or '')
    period.notes = request.form.get('notes', period.notes)
    
    db.session.commit()
    return jsonify({'success': True, 'message': '更新成功'})

# 编辑页面
@app.route('/edit_page/<int:period_id>', methods=['GET', 'POST'])
@login_required
def edit_page(period_id):
    period = Period.query.get_or_404(period_id)
    if period.user_id != current_user.id:
        flash('无权修改此记录', 'error')
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        period.start_date = datetime.strptime(request.form.get('start_date'), '%Y-%m-%d').date()
        period.end_date = datetime.strptime(request.form.get('end_date'), '%Y-%m-%d').date()
        period.flow = request.form.get('flow', period.flow)
        period.pain_level = int(request.form.get('pain_level', 0))
        period.symptoms = request.form.get('symptoms', '')
        period.notes = request.form.get('notes', '')
        db.session.commit()
        flash('记录更新成功！', 'success')
        return redirect(url_for('index'))
    
    return render_template('edit_period.html', period=period)

# 批量删除记录
@app.route('/batch_delete', methods=['POST'])
@login_required
def batch_delete_period():
    ids = request.form.get('ids', '').split(',')
    deleted_count = 0
    for id_str in ids:
        if id_str.strip():
            try:
                pid = int(id_str.strip())
                period = Period.query.filter_by(id=pid, user_id=current_user.id).first()
                if period:
                    db.session.delete(period)
                    deleted_count += 1
            except:
                pass
    db.session.commit()
    flash(f'已删除 {deleted_count} 条记录', 'success')
    return redirect(url_for('index'))

# API: 获取当前用户数据
@app.route('/api/data')
@login_required
def api_data():
    periods = Period.query.filter_by(user_id=current_user.id).all()
    data = []
    for p in periods:
        data.append({
            'id': p.id,
            'start_date': p.start_date.strftime('%Y-%m-%d'),
            'end_date': p.end_date.strftime('%Y-%m-%d'),
            'flow': p.flow,
            'pain_level': p.pain_level or 0,
            'symptoms': p.symptoms or '',
            'notes': p.notes
        })
    return jsonify(data)

# 数据导出
@app.route('/export')
@login_required
def export_data():
    import json
    format_type = request.args.get('format', 'xlsx')
    
    # 获取所有数据
    periods = Period.query.filter_by(user_id=current_user.id).all()
    studies = StudyRecord.query.filter_by(user_id=current_user.id).all()
    works = WorkRecord.query.filter_by(user_id=current_user.id).all()
    
    if format_type == 'xlsx':
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        
        wb = Workbook()
        
        # 月经记录Sheet
        ws1 = wb.active
        ws1.title = "月经记录"
        headers1 = ['开始日期', '结束日期', '经量', '疼痛等级', '症状', '备注']
        ws1.append(headers1)
        for p in periods:
            flow_name = {'light': '少量', 'medium': '中等', 'heavy': '大量'}.get(p.flow, p.flow)
            ws1.append([
                p.start_date.strftime('%Y-%m-%d'),
                p.end_date.strftime('%Y-%m-%d'),
                flow_name,
                p.pain_level or 0,
                p.symptoms or '',
                p.notes or ''
            ])
        
        # 学习记录Sheet
        ws2 = wb.create_sheet("学习记录")
        headers2 = ['日期', '科目', '时长(分钟)', '计划', '备注']
        ws2.append(headers2)
        for s in studies:
            ws2.append([
                s.date.strftime('%Y-%m-%d'),
                s.subject,
                s.duration,
                s.plan or '',
                s.notes or ''
            ])
        
        # 工作记录Sheet
        ws3 = wb.create_sheet("工作记录")
        headers3 = ['日期', '任务', '时长(分钟)', '上班时间', '下班时间', '加班时长(分钟)', '备注']
        ws3.append(headers3)
        for w in works:
            ws3.append([
                w.date.strftime('%Y-%m-%d'),
                w.task or '',
                w.task_duration,
                w.work_start.strftime('%H:%M') if w.work_start else '',
                w.work_end.strftime('%H:%M') if w.work_end else '',
                w.overtime,
                w.notes or ''
            ])
        
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = app.response_class(
            response=output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename=日历记录_{datetime.now().strftime("%Y%m%d")}.xlsx'}
        )
        return response
    
    # JSON格式
    export_data = {
        'export_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'user': current_user.username,
        'periods': [{
            'start_date': p.start_date.strftime('%Y-%m-%d'),
            'end_date': p.end_date.strftime('%Y-%m-%d'),
            'flow': p.flow,
            'pain_level': p.pain_level or 0,
            'symptoms': p.symptoms or '',
            'notes': p.notes
        } for p in periods],
        'studies': [{
            'date': s.date.strftime('%Y-%m-%d'),
            'subject': s.subject,
            'duration': s.duration,
            'plan': s.plan or '',
            'notes': s.notes or ''
        } for s in studies],
        'works': [{
            'date': w.date.strftime('%Y-%m-%d'),
            'task': w.task or '',
            'duration': w.task_duration,
            'work_start': w.work_start.strftime('%H:%M') if w.work_start else '',
            'work_end': w.work_end.strftime('%H:%M') if w.work_end else '',
            'overtime': w.overtime,
            'notes': w.notes or ''
        } for w in works]
    }
    
    if format_type == 'json':
        response = app.response_class(
            response=json.dumps(export_data, ensure_ascii=False, indent=2),
            mimetype='application/json',
            headers={'Content-Disposition': f'attachment; filename=calendar_export_{datetime.now().strftime("%Y%m%d")}.json'}
        )
        return response
    
    return jsonify({'success': False, 'message': '不支持的格式'})
    
    return jsonify({'success': False, 'message': '不支持的格式'})

# 排卵期提醒计算
@app.route('/api/ovulation')
@login_required
def get_ovulation_info():
    """获取排卵期信息"""
    periods = Period.query.filter_by(user_id=current_user.id).order_by(Period.start_date.desc()).limit(10).all()
    
    if len(periods) < 2:
        return jsonify({'has_data': False, 'message': '需要至少2条经期记录'})
    
    # 计算平均周期
    cycles = []
    for i in range(1, len(periods)):
        cycle = (periods[i-1].start_date - periods[i].start_date).days
        if 20 <= cycle <= 45:  # 合理周期范围
            cycles.append(cycle)
    
    if not cycles:
        return jsonify({'has_data': False, 'message': '周期数据不足以计算'})
    
    avg_cycle = sum(cycles) / len(cycles)
    last_period = periods[0]
    
    # 排卵日 = 下次经期开始日 - 14天
    # 假设周期为28天，排卵日在经期开始后第14天
    next_period_start = last_period.start_date + timedelta(days=int(avg_cycle))
    ovulation_date = next_period_start - timedelta(days=14)
    
    # 排卵期 = 排卵日前5天 + 后4天
    fertile_start = ovulation_date - timedelta(days=5)
    fertile_end = ovulation_date + timedelta(days=4)
    
    today = datetime.now().date()
    
    return jsonify({
        'has_data': True,
        'avg_cycle': int(avg_cycle),
        'last_period': last_period.start_date.strftime('%Y-%m-%d'),
        'next_period': next_period_start.strftime('%Y-%m-%d'),
        'ovulation_date': ovulation_date.strftime('%Y-%m-%d'),
        'fertile_start': fertile_start.strftime('%Y-%m-%d'),
        'fertile_end': fertile_end.strftime('%Y-%m-%d'),
        'days_until_ovulation': (ovulation_date - today).days,
        'is_fertile': fertile_start <= today <= fertile_end,
        'is_ovulation_day': today == ovulation_date
    })

# 检查邮箱是否已注册 (AJAX)
@app.route('/api/check_email')
def check_email():
    email = request.args.get('email')
    user = User.query.filter_by(email=email).first()
    return jsonify({'exists': user is not None})

# ==================== AI 功能 ====================

def call_deepseek(prompt, system_prompt=None):
    """调用 DeepSeek API"""
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {DEEPSEEK_API_KEY}"
    }
    
    messages = []
    if system_prompt:
        messages.append({"role": "system", "content": system_prompt})
    messages.append({"role": "user", "content": prompt})
    
    data = {
        "model": "deepseek-chat",
        "messages": messages,
        "temperature": 0.7,
        "max_tokens": 1000
    }
    
    try:
        response = requests.post(DEEPSEEK_API_URL, headers=headers, json=data, timeout=60)
        result = response.json()
        return result['choices'][0]['message']['content']
    except Exception as e:
        return f"抱歉，AI服务暂时不可用: {str(e)}"

# AI 预测经期
@app.route('/api/ai/predict')
@login_required
def ai_predict_period():
    """基于历史数据预测下次经期"""
    periods = Period.query.filter_by(user_id=current_user.id).order_by(Period.start_date.desc()).limit(10).all()
    
    if len(periods) < 2:
        return jsonify({
            'success': False,
            'message': '需要至少2条经期记录才能进行预测'
        })
    
    # 计算平均周期
    cycles = []
    for i in range(len(periods) - 1):
        cycle = (periods[i].start_date - periods[i+1].start_date).days
        cycles.append(cycle)
    
    avg_cycle = sum(cycles) / len(cycles)
    last_period = periods[0]
    predicted_date = last_period.start_date + timedelta(days=int(avg_cycle))
    
    # 准备历史数据给AI分析
    history_text = "\n".join([
        f"第{i+1}次: {p.start_date} 到 {p.end_date}, 经量: {p.flow}"
        for i, p in enumerate(periods[:6])
    ])
    
    system_prompt = """你是一个专业的经期健康助手。根据用户的历史经期记录，分析规律并给出预测和健康建议。
请用温暖、关心的语气回复，回复使用中文。"""
    
    prompt = f"""用户的历史经期记录：
{history_text}

平均经期周期: {avg_cycle:.1f} 天
最近一次经期开始日期: {last_period.start_date}
预测下次经期开始日期: {predicted_date}

请分析用户的经期规律，并给出：
1. 预测的下次经期日期
2. 经期健康建议（如饮食、运动、休息等）
3. 注意事项

请用简洁友好的方式回复。"""
    
    ai_response = call_deepseek(prompt, system_prompt)
    
    return jsonify({
        'success': True,
        'predicted_date': predicted_date.strftime('%Y-%m-%d'),
        'avg_cycle': round(avg_cycle, 1),
        'ai_analysis': ai_response
    })

# AI 健康建议
@app.route('/api/ai/advice', methods=['POST'])
@login_required
def ai_get_advice():
    """获取个性化健康建议"""
    data = request.get_json()
    user_input = data.get('question', '')
    
    # 获取用户最近的经期数据
    periods = Period.query.filter_by(user_id=current_user.id).order_by(Period.start_date.desc()).limit(3).all()
    
    history_text = "暂无记录"
    if periods:
        history_text = "\n".join([
            f"- {p.start_date} 到 {p.end_date}, 经量: {p.flow}"
            for p in periods
        ])
    
    system_prompt = """你是一个专业、温暖的经期健康顾问。请根据用户的问题和经期历史记录，给出专业而贴心的建议。
注意：如果用户询问的是医疗相关问题，请提醒用户咨询专业医生。"""
    
    prompt = f"""用户最近3次经期记录：
{history_text}

用户问题: {user_input}

请根据用户的经期历史和当前问题，给出适合的建议。如果用户没有提供具体问题，请给出通用的经期健康建议。"""
    
    ai_response = call_deepseek(prompt, system_prompt)
    
    return jsonify({
        'success': True,
        'advice': ai_response
    })

# AI 聊天助手
@app.route('/api/ai/chat', methods=['POST'])
@login_required
def ai_chat():
    """AI 聊天助手"""
    data = request.get_json()
    message = data.get('message', '')
    
    if not message:
        return jsonify({'success': False, 'message': '请输入内容'})
    
    # 获取用户历史记录作为上下文
    periods = Period.query.filter_by(user_id=current_user.id).order_by(Period.start_date.desc()).limit(5).all()
    
    history_text = "暂无记录"
    if periods:
        history_text = "\n".join([
            f"- {p.start_date} 到 {p.end_date}, 经量: {p.flow}"
            for p in periods
        ])
    
    system_prompt = """你是一个温柔、专业、亲切的经期健康助手"小经"。你了解经期健康知识，能够回答用户关于经期的问题。
请用轻松友好的语气聊天，适当使用emoji，回复控制在200字以内。"""
    
    prompt = f"""用户的经期历史记录：
{history_text}

用户说: {message}

请以"小经"的身份回复用户。"""
    
    ai_response = call_deepseek(prompt, system_prompt)
    
    return jsonify({
        'success': True,
        'reply': ai_response
    })

# 意图识别和提取
import re
from datetime import datetime, timedelta

def extract_date(text):
    """从文本中提取日期"""
    today = datetime.now().date()
    
    # 处理"今天"、"明天"、"后天"
    if '今天' in text:
        return today
    if '明天' in text:
        return today + timedelta(days=1)
    if '后天' in text:
        return today + timedelta(days=2)
    if '昨天' in text:
        return today - timedelta(days=1)
    if '前天' in text:
        return today - timedelta(days=2)
    
    # 处理"周一"~"周日"
    weekdays = {'周一': 0, '周二': 1, '周三': 2, '周四': 3, '周五': 4, '周六': 5, '周日': 6}
    for day, offset in weekdays.items():
        if day in text:
            current_weekday = today.weekday()
            days_to_add = (offset - current_weekday + 7) % 7
            if days_to_add == 0:
                days_to_add = 7
            return today + timedelta(days=days_to_add)
    
    # 处理"X月X日"或"X/X"
    month_day_pattern = re.search(r'(\d{1,2})[月/](\d{1,2})', text)
    if month_day_pattern:
        month = int(month_day_pattern.group(1))
        day = int(month_day_pattern.group(2))
        year = today.year
        if month < today.month:
            year += 1
        try:
            return datetime(year, month, day).date()
        except:
            pass
    
    # 处理"X号"
    day_pattern = re.search(r'(\d{1,2})号', text)
    if day_pattern:
        day = int(day_pattern.group(1))
        month = today.month
        year = today.year
        if day < today.day:
            month += 1
            if month > 12:
                month = 1
                year += 1
        try:
            return datetime(year, month, day).date()
        except:
            pass
    
    return None

def extract_time(text):
    """从文本中提取时间"""
    time_pattern = re.search(r'(\d{1,2})[点时](\d{1,2})?分?', text)
    if time_pattern:
        hour = int(time_pattern.group(1))
        minute = int(time_pattern.group(2)) if time_pattern.group(2) else 0
        return f"{hour:02d}:{minute:02d}"
    
    am_pattern = re.search(r'上午(\d{1,2})[点时]', text)
    if am_pattern:
        hour = int(am_pattern.group(1))
        return f"{hour:02d}:00"
    
    pm_pattern = re.search(r'下午(\d{1,2})[点时]', text)
    if pm_pattern:
        hour = int(pm_pattern.group(1)) + 12
        if hour >= 24:
            hour -= 24
        return f"{hour:02d}:00"
    
    return None

def extract_duration(text):
    """从文本中提取时长（分钟）"""
    hour_pattern = re.search(r'(\d{1,2})小时', text)
    if hour_pattern:
        return int(hour_pattern.group(1)) * 60
    
    min_pattern = re.search(r'(\d{1,2})分钟', text)
    if min_pattern:
        return int(min_pattern.group(1))
    
    min_pattern2 = re.search(r'(\d{1,2})分', text)
    if min_pattern2:
        return int(min_pattern2.group(1))
    
    return None

def recognize_intent(text):
    """识别用户意图"""
    if any(kw in text for kw in ['月经', '经期', '大姨妈', '来月经']):
        return 'period'
    if any(kw in text for kw in ['学习', '看书', '上课', '读书', '自习']):
        return 'study'
    if any(kw in text for kw in ['工作', '上班', '开会', '会议', '任务']):
        return 'work'
    if any(kw in text for kw in ['查看', '查询', '看看', '有什么', '日程']):
        return 'query'
    return 'unknown'

def process_voice_command(message, user_id=1):
    """处理语音命令 - 自动提取和执行"""
    intent = recognize_intent(message)
    date = extract_date(message)
    time = extract_time(message)
    duration = extract_duration(message)
    
    extracted_info = {
        'intent': intent,
        'date': date.strftime('%Y-%m-%d') if date else None,
        'time': time,
        'duration': duration,
        'original_message': message
    }
    
    if intent == 'period':
        return handle_period_voice(message, extracted_info, user_id)
    elif intent == 'study':
        return handle_study_voice(message, extracted_info, user_id)
    elif intent == 'work':
        return handle_work_voice(message, extracted_info, user_id)
    elif intent == 'query':
        return handle_query_voice(message, user_id)
    else:
        return general_voice_reply(message)

def handle_period_voice(message, info, user_id):
    """处理月经记录语音"""
    flow = 'medium'
    if any(kw in message for kw in ['量多', '很多', '大量']):
        flow = 'heavy'
    elif any(kw in message for kw in ['量少', '很少', '少量']):
        flow = 'light'
    
    pain_level = 0
    pain_match = re.search(r'(\d+)分?疼痛?', message)
    if pain_match:
        pain_level = int(pain_match.group(1))
    elif '很痛' in message:
        pain_level = 7
    elif '痛' in message:
        pain_level = 5
    
    symptoms = []
    symptom_keywords = {'肚子疼': '腹痛', '腰疼': '腰痛', '疲劳': '疲劳', 
                       '困': '困倦', '长痘': '长痘'}
    for kw, symptom in symptom_keywords.items():
        if kw in message:
            symptoms.append(symptom)
    
    date_str = info.get('date')
    if not date_str:
        return {
            'success': True,
            'reply': "好的，要帮你记录月经呀。请告诉我开始日期是哪天？比如'今天'、'明天'或者'3月20号'~",
            'extracted_info': info
        }
    
    # 转换为date对象
    date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
    
    existing = Period.query.filter_by(user_id=user_id, start_date=date_obj).first()
    if existing:
        return {'success': True, 'reply': f"你{date_str}的月经记录已经有了。要修改吗？", 'record_id': existing.id}
    
    period = Period(user_id=user_id, start_date=date_obj, end_date=date_obj,
                   flow=flow, pain_level=pain_level, symptoms=','.join(symptoms) if symptoms else '')
    db.session.add(period)
    db.session.commit()
    
    return {'success': True, 'reply': f"好啦，已记录！\n📅 {date_str}\n💧 经量：{flow}\n😖 疼痛：{pain_level}/10", 'record_id': period.id}

def handle_study_voice(message, info, user_id):
    """处理学习记录语音"""
    subject = '学习'
    subjects = {'英语': '英语', '数学': '数学', '语文': '语文', '物理': '物理', '化学': '化学'}
    for kw, sub in subjects.items():
        if kw in message:
            subject = sub
            break
    
    date_str = info.get('date') or datetime.now().date().strftime('%Y-%m-%d')
    # 转换为date对象
    date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
    duration = info.get('duration') or 60
    
    record = StudyRecord(user_id=user_id, date=date_obj, subject=subject,
                         duration=duration, notes=f"语音记录")
    db.session.add(record)
    db.session.commit()
    
    return {'success': True, 'reply': f"学习记录已添加！\n📚 {subject}\n📅 {date_str}\n⏰ {duration}分钟", 'record_id': record.id}

def handle_work_voice(message, info, user_id):
    """处理工作记录语音"""
    task = "工作任务"
    task_keywords = {'开会': '开会', '会议': '会议', '上班': '上班', '项目': '项目', '任务': '任务'}
    for kw, t in task_keywords.items():
        if kw in message:
            task = t
            break
    
    date_str = info.get('date') or datetime.now().date().strftime('%Y-%m-%d')
    # 转换为date对象
    date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
    duration = info.get('duration') or 60
    
    record = WorkRecord(user_id=user_id, date=date_obj, task=task,
                       task_duration=duration, notes=f"语音记录")
    db.session.add(record)
    db.session.commit()
    
    return {'success': True, 'reply': f"工作记录已添加！\n📋 {task}\n📅 {date_str}\n⏰ {duration}分钟", 'record_id': record.id}

def handle_query_voice(message, user_id):
    """处理查询语音"""
    today = datetime.now().date()
    
    if '今天' in message:
        periods = Period.query.filter_by(user_id=user_id, start_date=today).all()
        studies = StudyRecord.query.filter_by(user_id=user_id, date=today).all()
        works = WorkRecord.query.filter_by(user_id=user_id, date=today).all()
        
        reply = f"📅 今天的日程：\n"
        if periods: reply += f"🌸 月经记录：{len(periods)}条\n"
        if studies: reply += f"📚 学习：{len(studies)}条\n"
        if works: reply += f"💼 工作：{len(works)}条\n"
        if not periods and not studies and not works: reply += "暂无记录"
        return {'success': True, 'reply': reply}
    
    if '本月' in message:
        periods = Period.query.filter_by(user_id=user_id).all()
        return {'success': True, 'reply': f"本月你有 {len(periods)} 条月经记录"}
    
    return {'success': True, 'reply': "可以说'查看今天日程'或'查看本月记录'"}

def general_voice_reply(message):
    """通用对话"""
    system_prompt = """你是一个智能日历助手，用友好简洁的语言回复。适当用emoji。"""
    ai_response = call_deepseek(message, system_prompt)
    return {'success': True, 'reply': ai_response}

# 语音对话接口
@app.route('/api/voice/chat', methods=['POST'])
def voice_chat():
    """小程序语音对话接口 - 带意图识别"""
    data = request.get_json()
    message = data.get('message', '')
    user_id = data.get('user_id', 1)
    
    if not message:
        return jsonify({'success': False, 'message': '请输入内容'})
    
    result = process_voice_command(message, user_id)
    return jsonify(result)

@app.route('/api/voice/info', methods=['GET'])
def get_voice_info():
    return jsonify({'success': True, 'message': '意图识别接口正常'})
@app.route('/api/ai/report')
@login_required
def ai_generate_report():
    """生成经期健康报告"""
    periods = Period.query.filter_by(user_id=current_user.id).order_by(Period.start_date.desc()).limit(10).all()
    
    if not periods:
        return jsonify({
            'success': False,
            'message': '暂无经期记录'
        })
    
    # 计算统计数据
    total_records = len(periods)
    cycle_days = []
    period_days = []
    
    for i in range(len(periods) - 1):
        cycle_days.append((periods[i].start_date - periods[i+1].start_date).days)
    
    for p in periods:
        period_days.append((p.end_date - p.start_date).days + 1)
    
    avg_cycle = sum(cycle_days) / len(cycle_days) if cycle_days else 28
    avg_period = sum(period_days) / len(period_days)
    
    # 统计经量分布
    flow_count = {'light': 0, 'medium': 0, 'heavy': 0}
    for p in periods:
        if p.flow in flow_count:
            flow_count[p.flow] += 1
    
    history_text = "\n".join([
        f"第{i+1}次: {p.start_date} 开始，持续 {((p.end_date - p.start_date).days + 1)} 天，经量: {p.flow}"
        for i, p in enumerate(periods[:6])
    ])
    
    system_prompt = """你是一个专业的经期健康分析师。根据用户的经期数据，生成一份详细的健康报告。
请用清晰的结构化方式呈现，使用emoji让报告更生动。"""
    
    prompt = f"""请根据以下用户的经期数据，生成一份健康报告：

历史记录：
{history_text}

统计数据：
- 记录次数: {total_records} 次
- 平均周期: {avg_cycle:.1f} 天
- 平均经期时长: {avg_period:.1f} 天
- 经量分布: 少量{flow_count['light']}次, 中等{flow_count['medium']}次, 大量{flow_count['heavy']}次

请生成包括以下内容的健康报告：
1. 周期分析 - 你的经期规律是否正常
2. 健康评分 (1-10分)
3. 存在的问题和建议
4. 下次经期预测
5. 温馨小贴士

回复使用中文，报告结构清晰。"""
    
    ai_response = call_deepseek(prompt, system_prompt)
    
    return jsonify({
        'success': True,
        'report': ai_response,
        'stats': {
            'avg_cycle': round(avg_cycle, 1),
            'avg_period': round(avg_period, 1),
            'total_records': total_records
        }
    })

# ==================== 学习 AI 功能 ====================

# AI 学习计划生成
@app.route('/api/ai/study/plan', methods=['POST'])
@login_required
def ai_study_plan():
    """生成学习计划"""
    data = request.get_json()
    subject = data.get('subject', '')
    goal = data.get('goal', '')
    days = data.get('days', 7)
    
    if not subject:
        return jsonify({'success': False, 'message': '请输入学习科目'})
    
    prompt = f"""请为用户生成一个{days}天的{subject}学习计划。

用户目标: {goal if goal else '暂无具体目标'}

请生成一个超级详细的学习计划，必须包含以下内容：

📅 **第1天到第{days}天** 每天都需要有：
1. 📖 学习内容：具体要学习的知识点或章节
2. ⏱️ 学习时长：建议学习多少分钟
3. 🎯 今日目标：今天要达成什么
4. 💡 学习方法：用什么方法学习

请严格按照以下表格格式输出：
| 天数 | 学习内容 | 学习时长 | 今日目标 | 学习方法 |
|------|----------|----------|----------|----------|
| 第1天 | 内容 | X分钟 | 目标 | 方法 |
| 第2天 | 内容 | X分钟 | 目标 | 方法 |
...（依次列出所有{days}天）

使用emoji让计划更生动，回复使用中文。"""
    
    system_prompt = """你是一个专业、耐心的学习顾问。你擅长制定学习计划，帮助用户高效学习。"""
    
    ai_response = call_deepseek(prompt, system_prompt)
    
    return jsonify({
        'success': True,
        'plan': ai_response
    })

# AI 学习效率分析
@app.route('/api/ai/study/analyze')
@login_required
def ai_study_analyze():
    """分析学习效率"""
    records = StudyRecord.query.filter_by(user_id=current_user.id).order_by(StudyRecord.date.desc()).limit(20).all()
    
    if not records:
        return jsonify({
            'success': False,
            'message': '暂无学习记录，无法分析'
        })
    
    # 统计数据
    total_duration = sum(r.duration for r in records)
    subject_stats = {}
    for r in records:
        if r.subject not in subject_stats:
            subject_stats[r.subject] = {'count': 0, 'duration': 0}
        subject_stats[r.subject]['count'] += 1
        subject_stats[r.subject]['duration'] += r.duration
    
    # 按星期几统计
    weekday_stats = {i: 0 for i in range(7)}
    for r in records:
        weekday_stats[r.date.weekday()] += r.duration
    
    best_day = max(weekday_stats, key=weekday_stats.get)
    weekdays = ['周一', '周二', '周三', '周四', '周五', '周六', '周日']
    
    history_text = "\n".join([
        f"- {r.date}: {r.subject}, {r.duration}分钟"
        for r in records[:15]
    ])
    
    prompt = f"""请分析用户的学习数据，找出学习规律和效率问题：

学习记录：
{history_text}

统计数据：
- 总学习时长: {total_duration} 分钟
- 学习科目分布: {subject_stats}
- 一周各天学习时长: {weekdays[best_day]}学习时长最长

请给出：
1. 学习效率评分 (1-10分)
2. 学习习惯分析
3. 存在的问题
4. 改进建议
5. 最佳学习时间建议

回复使用中文，结构清晰。"""
    
    system_prompt = """你是一个专业的学习效率分析师。你擅长分析学习数据，发现学习规律和问题。"""
    
    ai_response = call_deepseek(prompt, system_prompt)
    
    return jsonify({
        'success': True,
        'analysis': ai_response,
        'stats': {
            'total_duration': total_duration,
            'best_day': weekdays[best_day],
            'subject_count': len(subject_stats)
        }
    })

# AI 错题本辅导
@app.route('/api/ai/study/wrongQuestion', methods=['POST'])
@login_required
def ai_wrong_question():
    """AI 错题本辅导"""
    data = request.get_json()
    question = data.get('question', '')
    subject = data.get('subject', '')
    user_answer = data.get('user_answer', '')
    
    if not question:
        return jsonify({'success': False, 'message': '请输入错题内容'})
    
    prompt = f"""用户有一道错题需要讲解：

科目: {subject if subject else '未指定'}
错题: {question}
用户的答案: {user_answer if user_answer else '暂无'}

请帮助用户：
1. 分析这道题的知识点
2. 找出用户的错误原因
3. 给出正确的解题思路
4. 讲解相关知识点
5. 出一道类似的练习题

请用通俗易懂的方式讲解，回复使用中文。"""
    
    system_prompt = """你是一个专业、耐心的学科辅导老师。你擅长讲解题目，分析错误原因，帮助学生掌握知识点。"""
    
    ai_response = call_deepseek(prompt, system_prompt)
    
    return jsonify({
        'success': True,
        'explanation': ai_response,
        'subject': subject,
        'question': question,
        'user_answer': user_answer
    })

# 保存错题到错题集
@app.route('/api/ai/study/saveWrongQuestion', methods=['POST'])
@login_required
def ai_save_wrong_question():
    """保存错题到错题集"""
    data = request.get_json()
    subject = data.get('subject', '')
    question = data.get('question', '')
    user_answer = data.get('user_answer', '')
    ai_explanation = data.get('ai_explanation', '')
    
    if not question:
        return jsonify({'success': False, 'message': '请输入错题内容'})
    
    # 提取知识点和练习题
    knowledge_points = ""
    practice_question = ""
    
    # 简单解析AI回复，提取知识点和练习题
    if "知识点" in ai_explanation:
        try:
            parts = ai_explanation.split("知识点")
            if len(parts) > 1:
                knowledge_part = parts[1].split("\n")[0]
                knowledge_points = knowledge_part.strip()
        except:
            pass
    
    if "练习题" in ai_explanation or "类似题" in ai_explanation:
        try:
            for line in ai_explanation.split("\n"):
                if "练习题" in line or "类似题" in line:
                    practice_question = line.split("：")[-1] if "：" in line else line
                    break
        except:
            pass
    
    wrong_q = WrongQuestion(
        user_id=current_user.id,
        subject=subject or '未分类',
        question=question,
        user_answer=user_answer,
        ai_explanation=ai_explanation,
        knowledge_points=knowledge_points,
        practice_question=practice_question
    )
    db.session.add(wrong_q)
    db.session.commit()
    
    return jsonify({
        'success': True,
        'message': '错题已保存到错题集',
        'id': wrong_q.id
    })

# 重新生成错题解答
@app.route('/api/ai/study/regenerateWrongQuestion', methods=['POST'])
@login_required
def ai_regenerate_wrong_question():
    """重新生成错题解答"""
    data = request.get_json()
    question = data.get('question', '')
    subject = data.get('subject', '')
    user_answer = data.get('user_answer', '')
    
    if not question:
        return jsonify({'success': False, 'message': '请输入错题内容'})
    
    prompt = f"""用户有一道错题需要讲解，请用不同的方式解答：

科目: {subject if subject else '未指定'}
错题: {question}
用户的答案: {user_answer if user_answer else '暂无'}

请用不同的解题思路和方法来讲解：
1. 分析这道题涉及的知识点
2. 详细讲解正确的解题方法
3. 指出常见错误原因
4. 总结解题技巧
5.出一道不同类型的练习题

请用通俗易懂、生动有趣的方式讲解，回复使用中文。"""
    
    system_prompt = """你是一个专业、耐心的学科辅导老师。你擅长用不同的方法讲解题目，帮助学生真正理解。"""
    
    ai_response = call_deepseek(prompt, system_prompt)
    
    return jsonify({
        'success': True,
        'explanation': ai_response
    })

# 获取错题集列表
@app.route('/api/ai/study/wrongQuestions')
@login_required
def get_wrong_questions():
    """获取错题集列表"""
    subject = request.args.get('subject', '')
    
    query = WrongQuestion.query.filter_by(user_id=current_user.id)
    if subject:
        query = query.filter_by(subject=subject)
    
    questions = query.order_by(WrongQuestion.created_at.desc()).all()
    
    data = []
    for q in questions:
        data.append({
            'id': q.id,
            'subject': q.subject,
            'question': q.question,
            'user_answer': q.user_answer,
            'ai_explanation': q.ai_explanation,
            'knowledge_points': q.knowledge_points,
            'practice_question': q.practice_question,
            'created_at': q.created_at.strftime('%Y-%m-%d %H:%M')
        })
    
    return jsonify({'success': True, 'questions': data})

import re
from datetime import datetime as dt

# AI 接受学习计划
@app.route('/api/ai/study/acceptPlan', methods=['POST'])
@login_required
def ai_accept_study_plan():
    """接受AI生成的学习计划，批量添加学习记录"""
    data = request.get_json()
    plan_text = data.get('plan', '')
    subject = data.get('subject', '')
    
    if not plan_text:
        return jsonify({'success': False, 'message': '请先生成学习计划'})
    
    # 解析计划内容，提取每天的学习内容
    # 匹配类似 "第1天" 或 "Day 1" 或 "1." 的模式
    pattern = r'(?:第\s*(\d+)\s*天|Day\s*(\d+)|(\d+)\.)'
    matches = re.findall(pattern, plan_text)
    
    if not matches:
        return jsonify({'success': False, 'message': '无法解析学习计划格式'})
    
    added_count = 0
    today = dt.now().date()
    
    # 解析计划内容，提取每天的学习内容
    lines = plan_text.split('\n')
    daily_plans = []
    max_day = 0
    
    # 首先找出总天数
    for line in lines:
        day_match = re.search(r'第\s*(\d+)\s*天', line)
        if day_match:
            day_num = int(day_match.group(1))
            max_day = max(max_day, day_num)
    
    if max_day == 0:
        return jsonify({'success': False, 'message': '无法解析学习计划格式'})
    
    # 解析每天的内容
    for line in lines:
        day_match = re.search(r'第\s*(\d+)\s*天', line)
        if day_match:
            day_num = int(day_match.group(1))
            
            # 提取学习时长 - 查找表格中的分钟数
            duration = 60  # 默认60分钟
            minute_match = re.search(r'(\d+)\s*分钟', line)
            if minute_match:
                duration = int(minute_match.group(1))
            
            # 提取学习内容 - 查找 "| 内容 |" 格式
            content = ""
            parts = line.split('|')
            if len(parts) >= 3:
                # 第二列通常是学习内容
                content = parts[2].strip() if parts[2].strip() else f"第{day_num}天学习"
            else:
                # 尝试从文本中提取
                content = line.replace(f'第{day_num}天', '').strip()
                if not content:
                    content = f"第{day_num}天学习"
            
            # 提取学习目标
            goal = ""
            if len(parts) >= 5:
                goal = parts[4].strip()
            
            # 提取学习方法
            method = ""
            if len(parts) >= 6:
                method = parts[5].strip()
            
            # 组合学习内容
            plan_content = f"📖 {content}"
            if goal:
                plan_content += f"\n🎯 目标: {goal}"
            if method:
                plan_content += f"\n💡 方法: {method}"
            
            daily_plans.append({
                'day': day_num,
                'duration': duration,
                'content': plan_content
            })
    
    if not daily_plans:
        return jsonify({'success': False, 'message': '无法解析学习计划格式'})
    
    added_count = 0
    
    # 只添加不超过max_day天的记录
    for plan_info in daily_plans:
        if plan_info['day'] > max_day:
            continue
            
        # 计划日期从今天开始
        plan_date = today + timedelta(days=plan_info['day'] - 1)
        
        # 创建学习记录
        record = StudyRecord(
            user_id=current_user.id,
            date=plan_date,
            subject=subject or '学习',
            duration=plan_info['duration'],
            plan=plan_info['content'],
            notes='AI生成学习计划'
        )
        db.session.add(record)
        added_count += 1
    
    db.session.commit()
    
    return jsonify({
        'success': True,
        'added_count': added_count,
        'message': f'成功添加 {added_count} 条学习记录'
    })

# ==================== 工作 AI 功能 ====================

@app.route('/api/ai/work/plan', methods=['POST'])
@login_required
def ai_work_plan():
    """生成工作计划"""
    data = request.get_json()
    task = data.get('task', '')
    goal = data.get('goal', '')
    days = data.get('days', 7)
    
    if not task:
        return jsonify({'success': False, 'message': '请输入工作任务'})
    
    prompt = f"""请为用户生成一个{days}天的工作计划。

工作内容: {task}
工作目标: {goal if goal else '暂无具体目标'}

请生成一个详细的工作计划，必须包含以下内容：

📅 **第1天到第{days}天** 每天都需要有：
1. 📋 工作任务：具体要完成的工作
2. ⏱️ 预计时长：建议工作多少小时
3. 🎯 今日目标：今天要达成什么
4. 💡 工作方法：用什么方法提高效率

请严格按照以下表格格式输出：
| 天数 | 工作任务 | 预计时长 | 今日目标 | 工作方法 |
|------|----------|----------|----------|----------|
| 第1天 | 任务 | X小时 | 目标 | 方法 |
| 第2天 | 任务 | X小时 | 目标 | 方法 |
...（依次列出所有{days}天）

使用emoji让计划更生动，回复使用中文。"""
    
    system_prompt = """你是一个专业、高效的工作顾问。你擅长制定工作计划，帮助用户提高工作效率。"""
    
    ai_response = call_deepseek(prompt, system_prompt)
    
    return jsonify({
        'success': True,
        'plan': ai_response
    })

@app.route('/api/ai/work/analyze')
@login_required
def ai_work_analyze():
    """分析工作情况和加班数据"""
    records = WorkRecord.query.filter_by(user_id=current_user.id).order_by(WorkRecord.date.desc()).limit(20).all()
    
    if not records:
        return jsonify({'success': False, 'message': '暂无工作记录，无法分析'})
    
    total_task_duration = sum(r.task_duration for r in records)
    total_overtime = sum(r.overtime for r in records)
    overtime_days = sum(1 for r in records if r.overtime > 0)
    
    weekdays = ['周一', '周二', '周三', '周四', '周五', '周六', '周日']
    weekday_stats = {i: {'task': 0, 'overtime': 0} for i in range(7)}
    for r in records:
        weekday_stats[r.date.weekday()]['task'] += r.task_duration
        weekday_stats[r.date.weekday()]['overtime'] += r.overtime
    
    busiest_day = max(weekday_stats, key=lambda x: weekday_stats[x]['task'])
    
    history_text = "\n".join([
        f"- {r.date}: {r.task or '打卡'}, 任务{r.task_duration}分钟, 加班{r.overtime}分钟"
        for r in records[:15]
    ])
    
    prompt = f"""请分析用户的工作数据：

工作记录：
{history_text}

统计数据：
- 总任务时长: {total_task_duration} 分钟
- 总加班时长: {total_overtime} 分钟
- 加班天数: {overtime_days} 天
- 最忙工作日: {weekdays[busiest_day]}

请给出：
1. 工作效率评分 (1-10分)
2. 工作习惯分析
3. 加班问题诊断
4. 改进建议

回复使用中文，结构清晰。"""
    
    system_prompt = """你是一个专业的工作效率分析师。"""
    
    ai_response = call_deepseek(prompt, system_prompt)
    
    return jsonify({
        'success': True,
        'analysis': ai_response,
        'stats': {'total_task_duration': total_task_duration, 'total_overtime': total_overtime, 'busiest_day': weekdays[busiest_day]}
    })

@app.route('/api/ai/work/weeklyReport')
@login_required
def ai_weekly_report():
    """生成工作周报"""
    today = datetime.now().date()
    week_start = today - timedelta(days=today.weekday())
    week_end = week_start + timedelta(days=6)
    
    records = WorkRecord.query.filter(
        WorkRecord.user_id == current_user.id,
        WorkRecord.date >= week_start,
        WorkRecord.date <= week_end
    ).order_by(WorkRecord.date).all()
    
    if not records:
        return jsonify({'success': False, 'message': '本周暂无工作记录'})
    
    total_duration = sum(r.task_duration for r in records)
    total_overtime = sum(r.overtime for r in records)
    
    history_text = "\n".join([
        f"- {r.date}: {r.task or '打卡'}, {r.task_duration}分钟"
        for r in records
    ])
    
    prompt = f"""请根据用户本周的工作记录生成周报：

本周记录：
{history_text}

统计：总时长{total_duration}分钟，加班{total_overtime}分钟

请生成工作周报，包括：本周工作内容、数据统计、改进建议。"""
    
    ai_response = call_deepseek(prompt, "你是一个专业的工作周报助手。")
    
    return jsonify({
        'success': True,
        'report': ai_response,
        'stats': {'total_hours': round(total_duration/60, 1), 'overtime_hours': round(total_overtime/60, 1)}
    })

@app.route('/api/ai/work/acceptPlan', methods=['POST'])
@login_required
def ai_accept_work_plan():
    """接受工作计划"""
    data = request.get_json()
    plan_text = data.get('plan', '')
    task = data.get('task', '')
    
    if not plan_text:
        return jsonify({'success': False, 'message': '请先生成工作计划'})
    
    import re
    lines = plan_text.split('\n')
    daily_plans = []
    max_day = 0
    
    # 解析每天的计划内容
    for i, line in enumerate(lines):
        day_match = re.search(r'第\s*(\d+)\s*天', line)
        if day_match:
            day_num = int(day_match.group(1))
            max_day = max(max_day, day_num)
            
            duration = 60
            hour_match = re.search(r'(\d+)\s*小时', line)
            if hour_match:
                duration = int(hour_match.group(1)) * 60
            
            # 提取该行的内容作为备注
            note_content = line.strip()
            
            daily_plans.append({'day': day_num, 'duration': duration, 'note': note_content})
    
    if not daily_plans:
        return jsonify({'success': False, 'message': '无法解析计划'})
    
    added_count = 0
    today_date = datetime.now().date()
    
    for plan_info in daily_plans:
        plan_date = today_date + timedelta(days=plan_info['day'] - 1)
        # 把AI生成的具体计划内容放入备注
        note = f"AI工作计划 - {plan_info['note']}"
        record = WorkRecord(user_id=current_user.id, date=plan_date, task=task or '工作计划', task_duration=plan_info['duration'], notes=note)
        db.session.add(record)
        added_count += 1
    
    db.session.commit()
    
    return jsonify({'success': True, 'added_count': added_count})

# 情侣共享模式页面
@app.route('/partner')
@login_required
def partner_page():
    """情侣共享模式主页"""
    # 检查是否已有情侣
    partner = User.query.get(current_user.partner_id) if current_user.partner_id else None
    return render_template('partner.html', partner=partner)

# 绑定情侣 - 发起请求
@app.route('/partner/bind', methods=['GET', 'POST'])
@login_required
def partner_bind():
    """发起情侣绑定请求"""
    if request.method == 'POST':
        partner_username = request.form.get('username')
        partner = User.query.filter_by(username=partner_username).first()
        
        if not partner:
            flash('用户不存在！', 'error')
            return redirect(url_for('partner_bind'))
        
        if partner.id == current_user.id:
            flash('不能绑定自己！', 'error')
            return redirect(url_for('partner_bind'))
        
        if current_user.partner_id:
            flash('你已经有情侣了！', 'error')
            return redirect(url_for('partner_page'))
        
        # 检查是否已有待处理的请求
        existing_request = Friend.query.filter_by(
            user_id=current_user.id, 
            friend_id=partner.id, 
            relation_type='partner',
            status='pending'
        ).first()
        
        if existing_request:
            flash('已发送过请求，等待对方同意！', 'error')
            return redirect(url_for('partner_page'))
        
        # 检查对方是否已经有情侣
        if partner.partner_id:
            flash('对方已经有情侣了！', 'error')
            return redirect(url_for('partner_bind'))
        
        # 创建绑定请求
        friend_request = Friend(
            user_id=current_user.id,
            friend_id=partner.id,
            relation_type='partner',
            status='pending'
        )
        db.session.add(friend_request)
        db.session.commit()
        
        flash(f'已向 {partner.username} 发送绑定请求，等待对方同意！', 'success')
        return redirect(url_for('partner_page'))
    
    return render_template('partner_bind.html')

# 处理情侣绑定请求
@app.route('/partner/request/<int:request_id>/<action>')
@login_required
def partner_request(request_id, action):
    """处理情侣绑定请求（同意/拒绝）"""
    friend_request = Friend.query.get(request_id)
    
    if not friend_request or friend_request.friend_id != current_user.id:
        flash('请求不存在！', 'error')
        return redirect(url_for('partner_page'))
    
    if friend_request.relation_type != 'partner':
        flash('无效的请求！', 'error')
        return redirect(url_for('partner_page'))
    
    if action == 'accept':
        # 检查是否已有情侣
        if current_user.partner_id:
            flash('你已经有情侣了！', 'error')
            return redirect(url_for('partner_page'))
        
        # 检查对方是否已有情侣
        partner = User.query.get(friend_request.user_id)
        if partner.partner_id:
            flash('对方已经有情侣了！', 'error')
            friend_request.status = 'rejected'
            db.session.commit()
            return redirect(url_for('partner_page'))
        
        # 双向绑定
        current_user.partner_id = partner.id
        partner.partner_id = current_user.id
        
        # 更新请求状态
        friend_request.status = 'accepted'
        
        # 也添加对方的好友关系
        existing_friend = Friend.query.filter_by(
            user_id=friend_request.user_id,
            friend_id=current_user.id,
            relation_type='partner'
        ).first()
        
        if not existing_friend:
            reverse_friend = Friend(
                user_id=friend_request.user_id,
                friend_id=current_user.id,
                relation_type='partner',
                status='accepted'
            )
            db.session.add(reverse_friend)
        
        db.session.commit()
        
        flash(f'已同意 {partner.username} 的绑定请求！', 'success')
    
    elif action == 'reject':
        friend_request.status = 'rejected'
        db.session.commit()
        flash('已拒绝绑定请求', 'info')
    
    return redirect(url_for('partner_page'))

# 解绑情侣
@app.route('/partner/unbind', methods=['POST'])
@login_required
def partner_unbind():
    """解绑情侣"""
    if not current_user.partner_id:
        flash('你还没有绑定情侣！', 'error')
        return redirect(url_for('partner_page'))
    
    partner = User.query.get(current_user.partner_id)
    if partner:
        partner.partner_id = None
        # 删除相关的好友关系
        Friend.query.filter(
            ((Friend.user_id == current_user.id) & (Friend.friend_id == partner.id)) |
            ((Friend.user_id == partner.id) & (Friend.friend_id == current_user.id)),
            Friend.relation_type == 'partner'
        ).delete()
    
    current_user.partner_id = None
    db.session.commit()
    
    flash('已成功解绑情侣！', 'success')
    return redirect(url_for('partner_page'))

# 查看情侣的经期记录
@app.route('/partner/periods')
@login_required
def partner_periods():
    """查看情侣的经期记录"""
    if not current_user.partner_id:
        flash('请先绑定情侣！', 'error')
        return redirect(url_for('partner_page'))
    
    partner = User.query.get(current_user.partner_id)
    periods = Period.query.filter_by(user_id=partner.id).order_by(Period.start_date.desc()).all()
    
    return render_template('partner_periods.html', partner=partner, periods=periods)

# 查看情侣的学习记录
@app.route('/partner/study')
@login_required
def partner_study():
    """查看情侣的学习记录"""
    if not current_user.partner_id:
        flash('请先绑定情侣！', 'error')
        return redirect(url_for('partner_page'))
    
    partner = User.query.get(current_user.partner_id)
    records = StudyRecord.query.filter_by(user_id=partner.id).order_by(StudyRecord.date.desc()).all()
    
    return render_template('partner_study.html', partner=partner, records=records)

# 查看情侣的工作记录
@app.route('/partner/work')
@login_required
def partner_work():
    """查看情侣的工作记录"""
    if not current_user.partner_id:
        flash('请先绑定情侣！', 'error')
        return redirect(url_for('partner_page'))
    
    partner = User.query.get(current_user.partner_id)
    records = WorkRecord.query.filter_by(user_id=partner.id).order_by(WorkRecord.date.desc()).all()
    
    return render_template('partner_work.html', partner=partner, records=records)

# API: 获取情侣的经期数据（日历用）
@app.route('/partner/api/periods')
@login_required
def partner_api_periods():
    """获取情侣的经期数据"""
    if not current_user.partner_id:
        return jsonify({'success': False, 'message': '请先绑定情侣'})
    
    partner = User.query.get(current_user.partner_id)
    periods = Period.query.filter_by(user_id=partner.id).all()
    
    events = []
    flow_colors = {'light': '#FFB6C1', 'medium': '#FF69B4', 'heavy': '#FF1493'}
    for p in periods:
        events.append({
            'id': p.id,
            'title': f'🌸 {partner.username}经期',
            'start': p.start_date.isoformat(),
            'end': (p.end_date + timedelta(days=1)).isoformat(),
            'color': flow_colors.get(p.flow, '#FF69B4'),
            'extendedProps': {
                'flow': p.flow,
                'pain_level': p.pain_level,
                'symptoms': p.symptoms,
                'notes': p.notes
            }
        })
    
    return jsonify({'success': True, 'events': events})

# API: 获取情侣的学习数据
@app.route('/partner/api/study')
@login_required
def partner_api_study():
    """获取情侣的学习数据"""
    if not current_user.partner_id:
        return jsonify({'success': False, 'message': '请先绑定情侣'})
    
    partner = User.query.get(current_user.partner_id)
    records = StudyRecord.query.filter_by(user_id=partner.id).all()
    
    events = []
    for r in records:
        events.append({
            'id': r.id,
            'title': f'📚 {r.subject} - {r.duration}分钟',
            'start': r.date.isoformat(),
            'color': '#4CAF50',
            'extendedProps': {
                'subject': r.subject,
                'duration': r.duration,
                'plan': r.plan,
                'notes': r.notes
            }
        })
    
    return jsonify({'success': True, 'events': events})

# API: 获取情侣的工作数据
@app.route('/partner/api/work')
@login_required
def partner_api_work():
    """获取情侣的工作数据"""
    if not current_user.partner_id:
        return jsonify({'success': False, 'message': '请先绑定情侣'})
    
    partner = User.query.get(current_user.partner_id)
    records = WorkRecord.query.filter_by(user_id=partner.id).all()
    
    events = []
    for r in records:
        events.append({
            'id': r.id,
            'title': f'💼 {r.task or "工作"} - {r.task_duration}分钟',
            'start': r.date.isoformat(),
            'color': '#2196F3',
            'extendedProps': {
                'task': r.task,
                'duration': r.task_duration,
                'overtime': r.overtime,
                'notes': r.notes
            }
        })
    
    return jsonify({'success': True, 'events': events})

# ========== 社区模块路由 ==========

# 话题选项
TOPICS = {
    'health': {'name': '经期健康', 'color': '#ff6b9d', 'icon': '💊'},
    'study': {'name': '学习打卡', 'color': '#1890ff', 'icon': '📚'},
    'work': {'name': '职场交流', 'color': '#fa8c16', 'icon': '💼'},
    'mood': {'name': '心情日记', 'color': '#52c41a', 'icon': '💭'},
    'general': {'name': '闲聊', 'color': '#722ed1', 'icon': '💬'}
}

@app.route('/community')
@login_required
def community():
    """社区首页"""
    topic_filter = request.args.get('topic', 'all')
    
    query = Post.query
    
    if topic_filter == 'following':
        # 只看关注用户的帖子
        following_ids = [f.following_id for f in UserFollow.query.filter_by(follower_id=current_user.id).all()]
        following_ids.append(current_user.id)  # 也显示自己的帖子
        query = query.filter(Post.user_id.in_(following_ids))
    elif topic_filter != 'all':
        query = query.filter_by(topic=topic_filter)
    
    posts = query.order_by(Post.created_at.desc()).limit(50).all()
    
    # 获取热门帖子（按点赞数）
    hot_posts = Post.query.order_by(Post.likes_count.desc()).limit(5).all()
    
    return render_template('community.html', 
                         posts=posts, 
                         hot_posts=hot_posts,
                         topics=TOPICS,
                         current_topic=topic_filter,
                         user=current_user)

@app.route('/community/upload_image', methods=['POST'])
@login_required
def community_upload_image():
    """单独上传图片"""
    if 'image' not in request.files:
        return jsonify({'success': False, 'message': '没有图片'})
    
    file = request.files['image']
    if not file or not file.filename:
        return jsonify({'success': False, 'message': '没有选择图片'})
    
    if not allowed_file(file.filename):
        return jsonify({'success': False, 'message': '不支持的图片格式'})
    
    ext = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else 'jpg'
    filename = f"{current_user.id}_{int(time.time())}.{ext}"
    file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
    
    return jsonify({
        'success': True, 
        'filename': filename,
        'url': url_for('static', filename='uploads/' + filename)
    })

@app.route('/community/post/add', methods=['POST'])
@login_required
def community_add_post():
    """发布帖子"""
    content = request.form.get('content', '').strip()
    if not content:
        flash('内容不能为空', 'error')
        return redirect(url_for('community'))
    
    topic = request.form.get('topic', 'general')
    is_anonymous = request.form.get('is_anonymous') == 'on'
    
    # 获取图片文件名
    images_str = request.form.get('images', '')
    images = images_str.split(',') if images_str else []
    images = [x for x in images if x]  # 过滤空值
    
    post = Post(
        user_id=current_user.id,
        content=content,
        images=','.join(images) if images else '',
        topic=topic,
        is_anonymous=is_anonymous
    )
    db.session.add(post)
    db.session.commit()
    
    flash('发布成功！', 'success')
    return redirect(url_for('community'))

@app.route('/community/post/<int:post_id>/like', methods=['POST'])
@login_required
def community_like_post(post_id):
    """点赞帖子"""
    post = Post.query.get_or_404(post_id)
    
    # 检查是否已点赞
    existing_like = PostLike.query.filter_by(user_id=current_user.id, post_id=post_id).first()
    
    if existing_like:
        db.session.delete(existing_like)
        post.likes_count = max(0, post.likes_count - 1)
    else:
        like = PostLike(user_id=current_user.id, post_id=post_id)
        db.session.add(like)
        post.likes_count += 1
    
    db.session.commit()
    return jsonify({'success': True, 'likes_count': post.likes_count})

@app.route('/community/post/<int:post_id>/comment', methods=['POST'])
@login_required
def community_comment_post(post_id):
    """评论帖子"""
    content = request.form.get('content', '').strip()
    if not content:
        return jsonify({'success': False, 'message': '评论不能为空'})
    
    comment = PostComment(
        user_id=current_user.id,
        post_id=post_id,
        content=content
    )
    db.session.add(comment)
    
    post = Post.query.get(post_id)
    post.comments_count += 1
    
    db.session.commit()
    return jsonify({'success': True, 'message': '评论成功'})

@app.route('/community/post/<int:post_id>')
@login_required
def community_post_detail(post_id):
    """帖子详情"""
    post = Post.query.get_or_404(post_id)
    comments = PostComment.query.filter_by(post_id=post_id, parent_id=None).order_by(PostComment.created_at.desc()).all()
    
    return render_template('community_post.html', post=post, comments=comments, topics=TOPICS, user=current_user)

@app.route('/community/post/<int:post_id>/delete', methods=['POST'])
@login_required
def community_delete_post(post_id):
    """删除帖子"""
    post = Post.query.get_or_404(post_id)
    if post.user_id != current_user.id:
        return jsonify({'success': False, 'message': '无权删除'})
    
    # 删除相关点赞和评论
    PostLike.query.filter_by(post_id=post_id).delete()
    PostComment.query.filter_by(post_id=post_id).delete()
    db.session.delete(post)
    db.session.commit()
    
    return jsonify({'success': True, 'message': '删除成功'})

@app.route('/community/post/<int:post_id>/edit', methods=['POST'])
@login_required
def community_edit_post(post_id):
    """编辑帖子"""
    post = Post.query.get_or_404(post_id)
    if post.user_id != current_user.id:
        return jsonify({'success': False, 'message': '无权修改'})
    
    content = request.form.get('content', '').strip()
    if not content:
        return jsonify({'success': False, 'message': '内容不能为空'})
    
    topic = request.form.get('topic', post.topic)
    is_anonymous = request.form.get('is_anonymous') == 'on'
    
    post.content = content
    post.topic = topic
    post.is_anonymous = is_anonymous
    
    db.session.commit()
    return jsonify({'success': True, 'message': '更新成功'})

@app.route('/community/user/<int:user_id>')
@login_required
def community_user_profile(user_id):
    """用户主页"""
    user = User.query.get_or_404(user_id)
    posts = Post.query.filter_by(user_id=user_id).order_by(Post.created_at.desc()).all()
    
    # 检查是否已关注
    is_following = UserFollow.query.filter_by(follower_id=current_user.id, following_id=user_id).first() is not None
    followers_count = UserFollow.query.filter_by(following_id=user_id).count()
    following_count = UserFollow.query.filter_by(follower_id=user_id).count()
    
    return render_template('community_profile.html', 
                         profile_user=user, 
                         posts=posts,
                         is_following=is_following,
                         followers_count=followers_count,
                         following_count=following_count,
                         topics=TOPICS,
                         user=current_user)

@app.route('/community/follow/<int:user_id>', methods=['POST'])
@login_required
def community_follow_user(user_id):
    """关注/取消关注用户"""
    if user_id == current_user.id:
        return jsonify({'success': False, 'message': '不能关注自己'})
    
    existing = UserFollow.query.filter_by(follower_id=current_user.id, following_id=user_id).first()
    
    if existing:
        db.session.delete(existing)
        db.session.commit()
        return jsonify({'success': True, 'following': False})
    else:
        follow = UserFollow(follower_id=current_user.id, following_id=user_id)
        db.session.add(follow)
        db.session.commit()
        return jsonify({'success': True, 'following': True})

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    app.run(debug=True, host='0.0.0.0', port=5000)
